import asyncio
import datetime
import json
import re
import uuid
from pathlib import Path

import aiofiles
from typing import Annotated, Optional
from openpyxl import Workbook
from rich.console import Console

import typer

from regulation_scraper.dynamic import load_module
from regulation_scraper.schemas.regulation import RegulationUnit, RegulationSection
from regulation_scraper.scrapers.regulation_scraper import RegulationScraper

app = typer.Typer()
console = Console()


def _is_cloudflare_or_cookie_error(exc: RuntimeError) -> bool:
    text = str(exc).lower()
    return (
        "cloudflare challenge" in text
        or "cf-mitigated" in text
        or "cookie" in text and "vencid" in text
    )


@app.command()
def scrape(
    out: Annotated[
        Path,
        typer.Argument(
            exists=False,
            help="Output directory where scraped regulations will be written",
        ),
    ],
    state: Annotated[Optional[list[str]], typer.Option()] = None,
):
    states = state or []
    if not states:
        print("No states provided...")
        return

    for state in states:
        state = state.lower()
        mod = load_module(f"./scrapers/{state}.py")
        scraper: RegulationScraper = mod.SCRAPER

        try:
            with console.status(f"Scraping {state}..."):
                asyncio.run(run(state, out, scraper))
        except RuntimeError as exc:
            if _is_cloudflare_or_cookie_error(exc):
                console.print(
                    "[bold yellow]ES:[/bold yellow] Cookie vencida o challenge de Cloudflare detectado. "
                    "Actualiza cookie/headers y vuelve a correr."
                )
                console.print(
                    "[bold yellow]EN:[/bold yellow] Expired cookie or Cloudflare challenge detected. "
                    "Refresh cookie/headers and run again."
                )
                raise typer.Exit(code=2)
            raise


async def run(state: str, out: Path, scraper: RegulationScraper):
    uid = "-".join(
        [
            datetime.datetime.now().strftime("%Y-%m-%d_%H_%M_%S"),
            str(uuid.uuid4()),
        ]
    )

    async def _prep(_type: str) -> Path:
        _out = out / uid / _type / f"{state}.jsonl"
        await aiofiles.os.makedirs(_out.parent, exist_ok=True)
        return _out

    written = 0
    units_rows: list[dict] = []
    sections_rows: list[dict] = []

    async with (
        aiofiles.open(await _prep("reg_unit"), "w", encoding="utf-8") as unit,
        aiofiles.open(await _prep("reg_section"), "w", encoding="utf-8") as section,
    ):
        async for reg in scraper.regulations():
            line_json = reg.model_dump_json()
            if isinstance(reg, RegulationUnit):
                print(f"Regulatory Unit: {reg.external_reference_id}")
                await unit.write(line_json + "\n")
                units_rows.append(reg.model_dump(mode="json"))
                written += 1
            elif isinstance(reg, RegulationSection):
                print(f"Regulatory Section: {reg.external_reference_id}")
                await section.write(line_json + "\n")
                sections_rows.append(reg.model_dump(mode="json"))
                written += 1
            else:
                raise ValueError(f"Unexpected regulation type: {type(reg)}")

    if written == 0:
        raise RuntimeError(
            "No regulations were scraped. Possible expired cookie or Cloudflare challenge."
        )

    _write_excel_report(
        out_dir=out / uid,
        state=state,
        units_rows=units_rows,
        sections_rows=sections_rows,
    )


def _write_excel_report(
    out_dir: Path,
    state: str,
    units_rows: list[dict],
    sections_rows: list[dict],
) -> None:
    workbook = Workbook()
    _write_scraped_sheet(workbook.active, sections_rows, units_rows)

    excel_path = out_dir / f"{state}_scrape.xlsx"
    workbook.save(excel_path)

    full_content_workbook = Workbook()
    _write_sheet(full_content_workbook.active, "reg_unit", units_rows)
    _write_sheet(
        full_content_workbook.create_sheet("reg_section"),
        "reg_section",
        sections_rows,
    )
    full_content_path = out_dir / f"{state}_from_json_full_content.xlsx"
    full_content_workbook.save(full_content_path)

    _write_sections_content_excel(out_dir=out_dir, state=state, sections_rows=sections_rows)


def _write_scraped_sheet(worksheet, sections_rows: list[dict], units_rows: list[dict]) -> None:
    worksheet.title = "scraped"
    worksheet.append(["type", "url", "name", "external_reference_id", "content"])

    for row_type, rows in (("section", sections_rows), ("unit", units_rows)):
        for row in rows:
            worksheet.append(
                [
                    row_type,
                    row.get("link"),
                    row.get("name"),
                    row.get("external_reference_id"),
                    row.get("content"),
                ]
            )
            url_cell = worksheet.cell(row=worksheet.max_row, column=2)
            if isinstance(url_cell.value, str) and url_cell.value.startswith("http"):
                url_cell.hyperlink = url_cell.value
                url_cell.style = "Hyperlink"


def _write_sheet(worksheet, title: str, rows: list[dict]) -> None:
    worksheet.title = title
    if not rows:
        worksheet.append(["message"])
        worksheet.append(["No data"])
        return

    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    worksheet.append(headers)
    for row in rows:
        values = [_excel_cell_value(row.get(header)) for header in headers]
        worksheet.append(values)
        row_number = worksheet.max_row
        for column_number, header in enumerate(headers, start=1):
            if header != "link":
                continue
            cell = worksheet.cell(row=row_number, column=column_number)
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def _excel_cell_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _is_continue_placeholder(content: str) -> bool:
    if not content:
        return False
    return bool(
        re.search(
            r"\[\s*please click here to continue\s*\]\(",
            content,
            flags=re.IGNORECASE,
        )
    )


def _write_sections_content_excel(out_dir: Path, state: str, sections_rows: list[dict]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "section_content"

    headers = [
        "id",
        "external_reference_id",
        "name",
        "status",
        "link",
        "unit_id",
        "index",
        "content",
        "content_length",
        "is_continue_placeholder",
    ]
    worksheet.append(headers)

    for row in sections_rows:
        content = str(row.get("content") or "")
        worksheet.append(
            [
                row.get("id"),
                row.get("external_reference_id"),
                row.get("name"),
                row.get("status"),
                row.get("link"),
                row.get("unit_id"),
                row.get("index"),
                content,
                len(content),
                _is_continue_placeholder(content),
            ]
        )
        link_cell = worksheet.cell(row=worksheet.max_row, column=5)
        if isinstance(link_cell.value, str) and link_cell.value.startswith("http"):
            link_cell.hyperlink = link_cell.value
            link_cell.style = "Hyperlink"

    excel_path = out_dir / f"{state}_sections_content.xlsx"
    workbook.save(excel_path)
